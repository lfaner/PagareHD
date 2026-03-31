from flask import Flask, render_template, request
from flask import send_from_directory
from datetime import date, datetime
import os
from calculos import calcular_neto_pagare
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

app = Flask(__name__)


def _parse_float(valor: str, campo: str) -> float:
    if valor is None:
        raise ValueError(f"Falta el valor de '{campo}'")
    valor = valor.strip().replace(",", ".")
    if valor == "":
        raise ValueError(f"Falta el valor de '{campo}'")
    try:
        return float(valor)
    except ValueError as exc:
        raise ValueError(f"'{campo}' debe ser numérico") from exc


def _normalize_header(value: str) -> str:
    return (
        value.strip()
        .lower()
        .replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
        .replace("ñ", "n")
    )


def _parse_excel_pagares(file_storage) -> list[dict]:
    wb = load_workbook(filename=file_storage, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("El Excel está vacío")

    headers = [str(c or "").strip() for c in rows[0]]
    norm_headers = [_normalize_header(h) for h in headers]

    col_map = {
        "valor_nominal": {"valor nominal", "valor_nominal", "vn", "valor"},
        "fecha_vencimiento": {
            "fecha vencimiento",
            "fecha_vencimiento",
            "vencimiento",
            "fecha vto",
            "vto",
        },
        "tna_descuento": {"tna descuento", "tna_descuento", "tna", "tasa"},
    }

    indices = {}
    for key, aliases in col_map.items():
        for idx, h in enumerate(norm_headers):
            if h in aliases:
                indices[key] = idx
                break

    data_start = 1
    if len(indices) < 3:
        indices = {"valor_nominal": 0, "fecha_vencimiento": 1, "tna_descuento": 2}
        data_start = 0

    form_rows = []
    row_id = 1
    for row in rows[data_start:]:
        vn  = row[indices["valor_nominal"]]      if len(row) > indices["valor_nominal"]      else None
        fv  = row[indices["fecha_vencimiento"]]  if len(row) > indices["fecha_vencimiento"]  else None
        tna = row[indices["tna_descuento"]]      if len(row) > indices["tna_descuento"]      else None

        if vn in (None, "") and fv in (None, "") and tna in (None, ""):
            continue

        if isinstance(fv, (int, float)):
            try:
                fv = from_excel(fv).date()
            except Exception:
                fv = str(fv).strip()
        if isinstance(fv, datetime):
            fv = fv.date()
        if isinstance(fv, date):
            fv = fv.isoformat()
        elif fv is not None:
            fv = str(fv).strip()

        form_rows.append(
            {
                "row_id":            str(row_id),
                "valor_nominal":     "" if vn  is None else str(vn),
                "fecha_vencimiento": "" if fv  is None else fv,
                "tna_descuento":     "" if tna is None else str(tna),
            }
        )
        row_id += 1

    if not form_rows:
        raise ValueError("No se encontraron pagarés en el Excel")

    return form_rows


def _defaults_base():
    return {
        "fecha_operacion":  date.today().isoformat(),
        "plazo_operacion":  "T+0",
        "tna_arancel":      "1.5",
        "comision_pct":     "0.5",
        "tipo_cambio_bna":  "",
    }


def _empty_row():
    return {
        "row_id":            "1",
        "valor_nominal":     "",
        "fecha_vencimiento": "",
        "tna_descuento":     "",
    }


@app.route("/", methods=["GET", "POST"])
def index():
    resultado   = None
    error       = None
    defaults    = _defaults_base()
    form_rows   = [_empty_row()]
    next_row_id = 2

    if request.method == "POST":
        try:
            # ===== Datos generales =====
            fecha_operacion_str = request.form.get("fecha_operacion")
            plazo_operacion     = request.form.get("plazo_operacion")
            tna_arancel_str     = request.form.get("tna_arancel")
            comision_pct_str    = request.form.get("comision_pct")
            tipo_cambio_str     = request.form.get("tipo_cambio_bna")

            for key, val in [
                ("fecha_operacion", fecha_operacion_str),
                ("plazo_operacion", plazo_operacion),
                ("tna_arancel",     tna_arancel_str),
                ("comision_pct",    comision_pct_str),
                ("tipo_cambio_bna", tipo_cambio_str),
            ]:
                if val:
                    defaults[key] = val

            if not all([fecha_operacion_str, plazo_operacion,
                        tna_arancel_str, comision_pct_str, tipo_cambio_str]):
                raise ValueError("Faltan datos generales de la operación")

            fecha_operacion  = datetime.strptime(fecha_operacion_str, "%Y-%m-%d").date()
            tna_arancel      = _parse_float(tna_arancel_str,  "Arancel TNA")
            comision_pct     = _parse_float(comision_pct_str, "Comisión")
            tipo_cambio_bna  = _parse_float(tipo_cambio_str,  "Tipo de cambio BNA")

            if plazo_operacion not in ("T+0", "T+1"):
                raise ValueError("El plazo debe ser 'T+0' o 'T+1'")
            if tna_arancel < 0:
                raise ValueError("El arancel no puede ser negativo")
            if comision_pct < 0:
                raise ValueError("La comisión no puede ser negativa")
            if tipo_cambio_bna <= 0:
                raise ValueError("El tipo de cambio BNA debe ser mayor a 0")

            # ===== Datos por pagaré =====
            valores       = request.form.getlist("valor_nominal[]")
            vencimientos  = request.form.getlist("fecha_vencimiento[]")
            tasas         = request.form.getlist("tna_descuento[]")
            row_ids       = request.form.getlist("row_id[]")
            delete_row_id = request.form.get("delete_row_id")

            if not valores or not vencimientos or not tasas:
                raise ValueError("Debe cargar al menos un pagaré completo")

            if not row_ids:
                row_ids = [str(i) for i in range(1, len(valores) + 1)]

            filtered = [
                (vn, fv, tna, rid)
                for vn, fv, tna, rid in zip(valores, vencimientos, tasas, row_ids)
                if not (delete_row_id and rid == delete_row_id)
            ]

            form_rows = [
                {"row_id": rid, "valor_nominal": vn,
                 "fecha_vencimiento": fv, "tna_descuento": tna}
                for vn, fv, tna, rid in filtered
            ]

            if not form_rows:
                form_rows = [_empty_row()]
                next_row_id = 2
            else:
                try:
                    next_row_id = max(int(r["row_id"]) for r in form_rows) + 1
                except ValueError:
                    next_row_id = len(form_rows) + 1

            detalle = []
            errores = []
            totales = {
                "cantidad_pagares":    0,
                "valor_nominal_usd":   0,
                "ppv":                 0,
                "descuento_usd":       0,
                "valor_descontado_usd": 0,
                "arancel_usd":         0,
                "comision_usd":        0,
                "iva_usd":             0,
                "neto_usd":            0,
                "derechos_mercado_ars": 0,
                "iva_derechos_ars":    0,
                "iibb_ars":            0,
            }
            ppv_num = 0
            ppv_den = 0

            for idx, (vn, fv, tna, rid) in enumerate(filtered, start=1):
                if not vn and not fv and not tna:
                    continue

                if not vn or not fv or not tna:
                    errores.append(f"Fila {idx}: complete valor nominal, vencimiento y TNA")
                    continue

                try:
                    valor_nominal  = _parse_float(vn,  f"Valor nominal (fila {idx})")
                    tna_descuento  = _parse_float(tna, f"TNA descuento (fila {idx})")
                    if valor_nominal <= 0:
                        raise ValueError("El valor nominal debe ser mayor a 0")
                    if tna_descuento < 0:
                        raise ValueError("La TNA de descuento no puede ser negativa")
                    fecha_vencimiento = datetime.strptime(fv, "%Y-%m-%d").date()
                except ValueError as e:
                    errores.append(f"Fila {idx}: {e}")
                    continue

                try:
                    res = calcular_neto_pagare(
                        valor_nominal=valor_nominal,
                        fecha_operacion=fecha_operacion,
                        fecha_vencimiento=fecha_vencimiento,
                        plazo_operacion=plazo_operacion,
                        tna_descuento=tna_descuento,
                        tna_arancel=tna_arancel,
                        comision_pct=comision_pct,
                        tipo_cambio_bna=tipo_cambio_bna,
                    )
                except Exception as e:
                    errores.append(f"Fila {idx}: {e}")
                    continue

                ppv_dias      = (res["fecha_cobro"] - fecha_operacion).days
                res["ppv_dias"] = ppv_dias
                res["row_id"]   = rid
                detalle.append(res)

                for k in totales:
                    totales[k] += res.get(k, 0)
                totales["cantidad_pagares"] += 1
                ppv_num += valor_nominal * ppv_dias
                ppv_den += valor_nominal

            if errores and not detalle:
                raise ValueError(" | ".join(errores))
            if errores:
                error = " | ".join(errores)
            if not detalle:
                raise ValueError("No hay pagarés válidos para calcular")

            if ppv_den > 0:
                totales["ppv"] = round(ppv_num / ppv_den, 0)

            totales["total_ars"] = round(
                totales["derechos_mercado_ars"]
                + totales["iva_derechos_ars"]
                + totales["iibb_ars"],
                2,
            )

            resultado = {
                "detalle":        detalle,
                "totales":        totales,
                "tipo_cambio_bna": tipo_cambio_bna,
            }

        except Exception as e:
            error = str(e)

    return render_template(
        "index.html",
        resultado=resultado,
        error=error,
        defaults=defaults,
        form_rows=form_rows,
        next_row_id=next_row_id,
    )


@app.route("/cargar-excel", methods=["POST"])
def cargar_excel():
    error       = None
    resultado   = None
    defaults    = _defaults_base()
    form_rows   = [_empty_row()]
    next_row_id = 2

    # Preservar datos generales enviados junto con el Excel
    for key in defaults:
        val = request.form.get(key)
        if val:
            defaults[key] = val

    file = request.files.get("excel_file")
    if not file or file.filename == "":
        error = "Debe seleccionar un archivo Excel"
    else:
        try:
            form_rows = _parse_excel_pagares(file)
            try:
                next_row_id = max(int(r["row_id"]) for r in form_rows) + 1
            except ValueError:
                next_row_id = len(form_rows) + 1
        except Exception as exc:
            error = str(exc)

    return render_template(
        "index.html",
        resultado=resultado,
        error=error,
        defaults=defaults,
        form_rows=form_rows,
        next_row_id=next_row_id,
    )


@app.route("/modelo-excel")
def modelo_excel():
    return send_from_directory("static", "modelo_pagares.xlsx", as_attachment=True)


if __name__ == "__main__":
    port = int(os.getenv("PORT", "5000"))
    app.run(host="0.0.0.0", port=port, debug=True)
